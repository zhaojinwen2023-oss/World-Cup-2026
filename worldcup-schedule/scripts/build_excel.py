from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from schedule_utils import default_app_data_path, ensure_output_dir, load_json


FULL_HEADERS = [
    "比赛编号",
    "比赛阶段",
    "轮次",
    "小组",
    "比赛日期",
    "对阵双方",
    "城市",
    "球场",
    "当地开球时间",
    "北京时间",
    "美国东部时间",
    "美国西部时间",
    "北京时间是否凌晨",
    "比赛状态",
    "当前分钟",
    "比分",
    "点球比分",
    "胜者",
    "是否已结束",
    "最后更新时间",
    "是否重点想看",
    "备注",
]

FAVORITE_HEADERS = ["比赛编号", "比赛阶段", "对阵双方", "北京时间", "美国东部时间", "美国西部时间", "比赛状态", "比分", "城市 / 球场", "备注"]
STANDINGS_HEADERS = ["小组", "排名", "球队", "赛", "胜", "平", "负", "进球", "失球", "净胜球", "积分", "公平竞赛分", "出线状态"]
THIRD_HEADERS = ["最佳第三排名", "小组", "球队", "积分", "净胜球", "进球", "公平竞赛分", "出线状态"]
KNOCKOUT_HEADERS = ["比赛编号", "轮次", "主队来源", "客队来源", "主队", "客队", "比分", "点球", "状态", "胜者", "下一场", "下一槽位", "北京时间", "美国东部", "美国西部"]
RESULT_HEADERS = ["比赛编号", "日期", "阶段", "对阵双方", "状态", "分钟", "比分", "点球", "胜者", "开球时间 / 更新时间", "是否重点想看"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FAVORITE_FILL = PatternFill("solid", fgColor="FFF2CC")
QUALIFIED_FILL = PatternFill("solid", fgColor="E2F0D9")
POSSIBLE_FILL = PatternFill("solid", fgColor="DDEBF7")
ELIMINATED_FILL = PatternFill("solid", fgColor="E7E6E6")
LIVE_FILL = PatternFill("solid", fgColor="F8CBAD")
EARLY_FILL = PatternFill("solid", fgColor="FCE4D6")
WINNER_FILL = PatternFill("solid", fgColor="D9EAD3")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2E7"),
    right=Side(style="thin", color="D9E2E7"),
    top=Side(style="thin", color="D9E2E7"),
    bottom=Side(style="thin", color="D9E2E7"),
)


def penalty(match: dict) -> str:
    home = str(match.get("home_penalty_score", "")).strip()
    away = str(match.get("away_penalty_score", "")).strip()
    return f"{home}-{away}" if home and away else ""


def full_row(match: dict) -> list:
    return [
        match["match_id"],
        match["stage_label"],
        match.get("round", ""),
        match.get("group", ""),
        match["beijing_date"],
        match["matchup"],
        match["city"],
        match["stadium"],
        match["local_display"],
        match["beijing_display"],
        match["eastern_display"],
        match["western_display"],
        "是" if match["beijing_is_early_morning"] else "否",
        match["status_label"],
        match.get("minute", ""),
        match.get("score", ""),
        penalty(match),
        match.get("winner", ""),
        "是" if match["is_finished"] else "否",
        match.get("last_updated", ""),
        match["favorite_mark"],
        match.get("notes", ""),
    ]


def favorite_row(match: dict) -> list:
    return [
        match["match_id"],
        match["stage_label"],
        match["matchup"],
        match["beijing_display"],
        match["eastern_display"],
        match["western_display"],
        match["status_label"],
        match.get("score", ""),
        f"{match['city']} / {match['stadium']}",
        match.get("notes", ""),
    ]


def result_row(match: dict) -> list:
    live_text = match.get("last_updated") if match["is_finished"] or match["is_live"] else match["beijing_display"]
    return [
        match["match_id"],
        match["beijing_date"],
        match["stage_label"],
        match["matchup"],
        match["status_label"],
        match.get("minute", ""),
        match.get("score", ""),
        penalty(match),
        match.get("winner", ""),
        live_text,
        match["favorite_mark"],
    ]


def knockout_row(row: dict, match_lookup: dict[int, dict]) -> list:
    match = match_lookup.get(row["match_id"], {})
    return [
        row["match_id"],
        row.get("round", ""),
        row.get("home_source", ""),
        row.get("away_source", ""),
        row.get("home_team", ""),
        row.get("away_team", ""),
        match.get("score", ""),
        penalty(match),
        match.get("status_label", ""),
        row.get("winner", match.get("winner", "")),
        row.get("next_match_id") or "",
        row.get("next_slot", ""),
        match.get("beijing_display", ""),
        match.get("eastern_display", ""),
        match.get("western_display", ""),
    ]


def style_table(ws, headers: list[str], table_name: str) -> None:
    max_row = ws.max_row
    max_col = len(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        title = ws.title
        values = [cell.value for cell in row]
        if title in {"完整赛程", "比赛结果"} and "⭐" in values:
            for cell in row:
                cell.fill = FAVORITE_FILL
        if title == "完整赛程" and row[12].value == "是":
            row[12].fill = EARLY_FILL
            row[12].font = Font(color="9C0006", bold=True)
        if title in {"完整赛程", "比赛结果"} and row[13 if title == "完整赛程" else 4].value in {"进行中", "中场", "加时", "点球"}:
            row[13 if title == "完整赛程" else 4].fill = LIVE_FILL
            row[13 if title == "完整赛程" else 4].font = Font(color="9C0006", bold=True)
        if title == "小组积分榜":
            status = row[12].value
            if status == "qualified":
                for cell in row:
                    cell.fill = QUALIFIED_FILL
            elif status == "possible":
                row[12].fill = POSSIBLE_FILL
            elif status == "eliminated":
                for cell in row:
                    cell.fill = ELIMINATED_FILL
                    cell.font = Font(color="7F7F7F")
        if title == "最佳小组第三":
            rank = row[0].value
            if isinstance(rank, int) and rank <= 8:
                for cell in row:
                    cell.fill = QUALIFIED_FILL
            else:
                for cell in row:
                    cell.fill = ELIMINATED_FILL
        if title == "淘汰赛对阵" and row[9].value:
            row[9].fill = WINNER_FILL
            row[9].font = Font(bold=True, color="375623")

    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    ws.add_table(table)
    ws.auto_filter.ref = ref
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    autofit_columns(ws, max_col)


def autofit_columns(ws, max_col: int) -> None:
    for column_index in range(1, max_col + 1):
        values = [str(ws.cell(row=row, column=column_index).value or "") for row in range(1, ws.max_row + 1)]
        width = min(max(max(len(value) for value in values) + 2, 9), 44)
        ws.column_dimensions[get_column_letter(column_index)].width = width


def build_info_sheet(ws, app_data: dict) -> None:
    ws.title = "使用说明"
    ws.sheet_view.showGridLines = False
    rows = [
        ("2026 世界杯动态赛程说明", ""),
        ("App 数据生成时间", app_data.get("generated_at", "")),
        ("比分更新时间", app_data.get("live_last_updated", "")),
        ("积分榜更新时间", app_data.get("standings_last_updated", "")),
        ("淘汰赛更新时间", app_data.get("knockout_last_updated", "")),
        ("自动更新命令", "python scripts/update_live_data.py --source local"),
        ("API Key", "写入 .env 或环境变量，不要写进代码。"),
        ("积分榜限制", "当前自动排名实现积分、净胜球、进球数、公平竞赛分；相互战绩和抽签未完全实现。"),
        ("Excel 实时性", "Excel 是脚本生成的分享文件，不会自己联网刷新；更新数据后需要重新生成。"),
    ]
    for index, (title, body) in enumerate(rows, start=1):
        ws.cell(index, 1, title)
        ws.cell(index, 2, body)
        ws.cell(index, 1).font = Font(bold=True, color="1F4E5F")
        ws.cell(index, 1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(index, 2).alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110


def build_workbook(app_data_path: Path, output_path: Path) -> None:
    app_data = load_json(app_data_path, {})
    matches = app_data.get("matches", [])
    standings = app_data.get("standings", [])
    best_thirds = app_data.get("best_thirds", [])
    knockout = app_data.get("knockout", [])
    match_lookup = {match["match_id"]: match for match in matches}

    wb = Workbook()
    ws_full = wb.active
    ws_full.title = "完整赛程"
    ws_full.append(FULL_HEADERS)
    for match in matches:
        ws_full.append(full_row(match))
    style_table(ws_full, FULL_HEADERS, "FullSchedule")

    ws_fav = wb.create_sheet("重点比赛")
    ws_fav.append(FAVORITE_HEADERS)
    for match in [item for item in matches if item.get("is_favorite")]:
        ws_fav.append(favorite_row(match))
    style_table(ws_fav, FAVORITE_HEADERS, "FavoriteSchedule")

    ws_standings = wb.create_sheet("小组积分榜")
    ws_standings.append(STANDINGS_HEADERS)
    for row in standings:
        ws_standings.append([row["group"], row["rank"], row["team"], row["played"], row["won"], row["drawn"], row["lost"], row["goals_for"], row["goals_against"], row["goal_difference"], row["points"], row.get("fair_play_points", 0), row.get("qualified_status", "unknown")])
    style_table(ws_standings, STANDINGS_HEADERS, "GroupStandings")

    ws_thirds = wb.create_sheet("最佳小组第三")
    ws_thirds.append(THIRD_HEADERS)
    for row in best_thirds:
        ws_thirds.append([row.get("best_third_rank", ""), row["group"], row["team"], row["points"], row["goal_difference"], row["goals_for"], row.get("fair_play_points", 0), row.get("qualified_status", "unknown")])
    style_table(ws_thirds, THIRD_HEADERS, "BestThirds")

    ws_knockout = wb.create_sheet("淘汰赛对阵")
    ws_knockout.append(KNOCKOUT_HEADERS)
    for row in knockout:
        ws_knockout.append(knockout_row(row, match_lookup))
    style_table(ws_knockout, KNOCKOUT_HEADERS, "KnockoutBracket")

    ws_results = wb.create_sheet("比赛结果")
    ws_results.append(RESULT_HEADERS)
    for match in matches:
        ws_results.append(result_row(match))
    style_table(ws_results, RESULT_HEADERS, "MatchResults")

    ws_info = wb.create_sheet("使用说明")
    build_info_sheet(ws_info, app_data)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build dynamic worldcup schedule Excel workbook.")
    parser.add_argument("--app-data", type=Path, default=default_app_data_path(), help="app_data.json 路径")
    parser.add_argument("--output", type=Path, default=ensure_output_dir() / "worldcup_2026_schedule.xlsx", help="输出 .xlsx 路径")
    args = parser.parse_args()
    build_workbook(args.app_data, args.output)
    print(f"已生成 Excel: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
